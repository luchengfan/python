import xml.dom.minidom
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import *

def check_language_type():
    '''
    检测输入的语言是否在excel中
    '''
    for lang_type in get_lang_type_list:
        for excel_title_content in excel_title:
            if lang_type in excel_title_content:
                input_in_title.append(lang_type)

    input_not_in_title = [item for item in get_lang_type_list if item not in input_in_title]

    if len(input_not_in_title) != 0:
        print(input_not_in_title , "不在语言列表中，请确认是否输入错误")
        exit(0)

def get_language_type_number():
    '''
    将输入的语言转化为excel表中的列数
    '''
    lang_type_num.append(0)
    lang_type_num.append(1)
    for input_type in input_in_title:
        for i in range(0 , len(excel_title)):
            if input_type in excel_title[i]:
                if ((i != 0) and (i != 1)):
                    lang_type_num.append(i)

def get_output_file(excel_file):
    '''
    设置输出的文件
    '''
    excel_filename = excel_file.split(r'/')[-1] #获取excel的文件名
    excel_folder_path = excel_file.replace(excel_filename , '')  #获取excel文件夹路径

    if is_dir:
        get_language_type_file = 'get_language_type/' + excel_folder_path
    else:
        get_language_type_file = 'get_language_type/'

    if not os.path.isdir(get_language_type_file):
        os.makedirs(get_language_type_file)

    output_file = get_language_type_file + excel_filename

    return output_file

def get_language_type(excel_file):
    print(excel_file , "开始处理！")
    input_not_in_title = []
    table_list = []
    styleRedBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')  # 红色

    data = xlrd.open_workbook(excel_file , formatting_info=True)
    sheet_table = data.sheets()[0] #读取excel表中的sheet1

    lang_output_file = get_output_file(excel_file)

    wb=xlwt.Workbook()   #新建一个excel文件
    ws=wb.add_sheet('translate')   #新增一个表格，并且取名translate

    for col in range(sheet_table.ncols): #列
        excel_title.append(sheet_table.cell(0,col).value)

    check_language_type()
    get_language_type_number()

    for row in range(sheet_table.nrows): #行
        col_num = 0
        for col in range(sheet_table.ncols): #列
            if col in lang_type_num:
                if sheet_table.cell(row,col).value != "":
                    ws.write(row , col_num , sheet_table.cell(row,col).value)
                else:
                    ws.write(row , col_num , sheet_table.cell(row,col).value , styleRedBkg)
                col_num += 1
            wb.save(lang_output_file)

    print(excel_file , "处理完成！")

def get_file(excel_file_path):
    '''
    获取传参路径下的xls/xlsx文件
    返回当前文件夹及其子目录下所有的xls/xlsx文件(包括绝对路径)
    '''
    xml_type_list = [".xls" , ".xlsx"]
    xml_list = []

    for root, dirs, files in os.walk(excel_file_path):
        for file in files:
            for xml_type in xml_type_list:
                if file.endswith(xml_type):
                    filename = root + "/" + file #注意不要改动"/"
                    xml_list.append(filename)
    return xml_list

if __name__ == "__main__":
    get_excel_file = input("请输入处理的文件或文件夹：")
    get_lang_type = input("请输入要保留的语言(Excel第一行,默认会保留英语)，以'_'区分，例如：values-es_values-it_values-ar：")
    get_lang_type_list = get_lang_type.split("_")
    
    excel_title = []
    input_in_title = []
    lang_type_num = []

    is_dir = False

    if os.path.isdir(get_excel_file): #文件夹
        is_dir = True
        get_language_type_file = get_file(get_excel_file)
        if len(get_language_type_file) == 0:
            print ("当前路径下无xml文件")
            exit(0)
        else:
            for file_path in get_language_type_file:
                get_language_type(file_path)
            print ("文件已处理完成，请查看get_language_type文件夹！")
    elif os.path.isfile(get_excel_file): #文件
        if (get_excel_file.endswith(".xls") or get_excel_file.endswith(".xlsx")):
            is_dir = False
            get_language_type(get_excel_file)
        else:
            print ("请输入xml或者excel文件")
            exit(0)
    else:
        print ("请检查输入的路径是否存在")
